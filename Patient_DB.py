from tkinter import*
from tkinter import ttk
import random
import time
import datetime
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
import uuid



class Hospital:
    def __init__(self, root):
        self.root=root
        self.root.title("Patient Database")
        self.root.geometry("1445x822+0+0")

        #===================================Variables===================================
        self.PatientName=StringVar()
        self.Age=StringVar()
        self.PatientAddress=StringVar()
        self.DateofBirth=StringVar()
        self.DateofVisit=StringVar()
        self.Symptoms=StringVar()
        self.Diagnosis=StringVar()
        self.Medication=StringVar()
        self.WardName=StringVar()
        self.DoctorName=StringVar()


        lbltitle=Label(self.root,bd=20,relief=RIDGE,text="PATIENT DATABASE",fg="black",bg="white",font=("times new roman",50,"bold"))
        lbltitle.pack(side=TOP,fill=X)

        #====================================Dataframe==================================
        Dataframe=Frame(self.root,bd=20,relief=RIDGE)
        Dataframe.place(x=0,y=130,width=1435,height=400)

        DataframeMid=LabelFrame(self.root,bd=10,relief=RIDGE,padx=10,font=("times new roman",12,"bold"),text="Patient Information")
        DataframeMid.place(x=20,y=150,width=1400,height=360)

        #====================================Buttons Frame==================================
        Buttonframe=Frame(self.root,bd=20,relief=RIDGE)
        Buttonframe.place(x=0,y=530,width=1435,height=100)

        #====================================Details Frame==================================
        Detailframe=Frame(self.root,bd=20,relief=RIDGE)
        Detailframe.place(x=0,y=600,width=1435,height=220)

        #====================================DataframeMid===================================

        lblPatientname=Label(DataframeMid,font=("arial",16,"bold"),text="Patient Name",padx=4,pady=8)
        lblPatientname.grid(row=0,column=0,sticky=W)
        txtPatientname=Entry(DataframeMid,textvariable=self.PatientName,font=("arial",16,"bold"),width=35)
        txtPatientname.grid(row=0,column=1)

        lblAge=Label(DataframeMid,font=("arial",16,"bold"),text="Age",padx=4,pady=8)
        lblAge.grid(row=1,column=0,sticky=W)
        txtAge=Entry(DataframeMid,textvariable=self.Age,font=("arial",16,"bold"),width=35)
        txtAge.grid(row=1,column=1)

        lblPatientaddress=Label(DataframeMid,font=("arial",16,"bold"),text="Patient Address",padx=4,pady=8)
        lblPatientaddress.grid(row=2,column=0,sticky=W)
        txtPatientaddress=Entry(DataframeMid,textvariable=self.PatientAddress,font=("arial",16,"bold"),width=35)
        txtPatientaddress.grid(row=2,column=1)

        lblDOB=Label(DataframeMid,font=("arial",16,"bold"),text="Date of Birth",padx=4,pady=8)
        lblDOB.grid(row=3,column=0,sticky=W)
        txtDOB=Entry(DataframeMid,textvariable=self.DateofBirth,font=("arial",16,"bold"),width=35)
        txtDOB.grid(row=3,column=1)

        lblDateofVisit=Label(DataframeMid,font=("arial",16,"bold"),text="Date of Visit",padx=4,pady=8)
        lblDateofVisit.grid(row=4,column=0,sticky=W)
        txtDateofVisit=Entry(DataframeMid,textvariable=self.DateofVisit,font=("arial",16,"bold"),width=35)
        txtDateofVisit.grid(row=4,column=1)

        lblSymptoms=Label(DataframeMid,font=("arial",16,"bold"),text="Symptoms",padx=4,pady=8)
        lblSymptoms.grid(row=5,column=0,sticky=W)
        txtSymptoms=Entry(DataframeMid,textvariable=self.Symptoms,font=("arial",16,"bold"),width=35)
        txtSymptoms.grid(row=5,column=1)

        lblDiagnosis=Label(DataframeMid,font=("arial",16,"bold"),text="Diagnosis",padx=4,pady=8)
        lblDiagnosis.grid(row=6,column=0,sticky=W)
        txtDiagnosis=Entry(DataframeMid,textvariable=self.Diagnosis,font=("arial",16,"bold"),width=35)
        txtDiagnosis.grid(row=6,column=1)

        lblMedication=Label(DataframeMid,font=("arial",16,"bold"),text="Medication",padx=4,pady=8)
        lblMedication.grid(row=0,column=3,sticky=W)
        txtMedication=Entry(DataframeMid,textvariable=self.Medication,font=("arial",16,"bold"),width=35)
        txtMedication.grid(row=0,column=4)

        lblWardname=Label(DataframeMid,font=("arial",16,"bold"),text="Ward Name",padx=4,pady=8)
        lblWardname.grid(row=1,column=3,sticky=W)
        txtWardname=Entry(DataframeMid,textvariable=self.WardName,font=("arial",16,"bold"),width=35)
        txtWardname.grid(row=1,column=4)

        lblDoctorname=Label(DataframeMid,font=("arial",16,"bold"),text="Doctor Name",padx=4,pady=8)
        lblDoctorname.grid(row=2,column=3,sticky=W)
        txtDoctorname=Entry(DataframeMid,textvariable=self.DoctorName,font=("arial",16,"bold"),width=35)
        txtDoctorname.grid(row=2,column=4)

        #================================Button======================================
        btnAdd=Button(Buttonframe,command=self.Add,text="Add",fg="white",bg="black",font=("arial",12,"bold"),width=23,height=16,padx=2,pady=6)
        btnAdd.pack(side=LEFT)

        btnUpdate=Button(Buttonframe,command=self.Update,text="Update",bg="black",fg="white",font=("arial",12,"bold"),width=23,height=16,padx=2,pady=6)
        btnUpdate.pack(side=LEFT)

        btnDelete=Button(Buttonframe,text="Delete",command=self.Delete,bg="black",fg="white",font=("arial",12,"bold"),width=23,height=16,padx=2,pady=6)
        btnDelete.pack(side=LEFT)

        btnClear=Button(Buttonframe,command=self.Clear,text="Clear",bg="black",fg="white",font=("arial",12,"bold"),width=23,height=16,padx=2,pady=6)
        btnClear.pack(side=LEFT)

        btnExit=Button(Buttonframe,command=lambda:root.destroy(),text="Exit",bg="black",fg="white",font=("arial",12,"bold"),width=23,height=16,padx=2,pady=6)
        btnExit.pack(side=LEFT)

        #================================Table======================================
        #================================Scrollbar======================================
        scroll_x=ttk.Scrollbar(Detailframe,orient=HORIZONTAL)
        scroll_y=ttk.Scrollbar(Detailframe,orient=VERTICAL)
        self.hospital_table=ttk.Treeview(Detailframe,column=("patientname", "age","patientaddress","dateofbirth","dateofvisit","symptoms","diagnosis","medication","referenceid","wardname","doctorname"),xscrollcommand=scroll_x.set,yscrollcommand=scroll_y.set)
        scroll_x.pack(side=BOTTOM,fill=X)
        scroll_y.pack(side=RIGHT,fill=Y)

        scroll_x=ttk.Scrollbar(command=self.hospital_table.xview)
        scroll_y=ttk.Scrollbar(command=self.hospital_table.yview)

        s=ttk.Style()
        s.configure("Treeview.Heading",font=("arial",12,"bold"))

        self.hospital_table.heading("patientname", text="Patient Name")
        self.hospital_table.heading("age", text="Age")
        self.hospital_table.heading("patientaddress", text="Patient Address")
        self.hospital_table.heading("dateofbirth", text="Date of Birth")
        self.hospital_table.heading("dateofvisit", text="Date of Visit")
        self.hospital_table.heading("symptoms", text="Symptoms")
        self.hospital_table.heading("diagnosis", text="Diagnosis")
        self.hospital_table.heading("referenceid", text="Reference ID")
        self.hospital_table.heading("medication", text="Medication")
        self.hospital_table.heading("wardname", text="Ward Name")
        self.hospital_table.heading("doctorname", text="Doctor Name")

        self.hospital_table["show"]="headings"
        
        self.hospital_table.column("patientname", width=100)
        self.hospital_table.column("age",width=100)
        self.hospital_table.column("patientaddress", width=100)
        self.hospital_table.column("dateofbirth", width=100)
        self.hospital_table.column("dateofvisit", width=100)
        self.hospital_table.column("symptoms", width=100)
        self.hospital_table.column("diagnosis", width=100)
        self.hospital_table.column("referenceid", width=100)
        self.hospital_table.column("medication", width=100)
        self.hospital_table.column("wardname", width=100)
        self.hospital_table.column("doctorname", width=100)

        self.hospital_table.pack(fill=BOTH,expand=1)
        self.hospital_table.bind("<ButtonRelease-1>",self.get_cursor)
        
        file=pathlib.Path('HMDS_data.xlsx')
        if file.exists():
            pass
        else:
            file=Workbook()
            sheet=file.active
            sheet['A1']="Patient Name"
            sheet['B1']="Age"
            sheet['C1']="Patient Address"
            sheet['D1']="Date of Birth"
            sheet['E1']="Date of Visit"
            sheet['F1']="Symptoms"
            sheet['G1']="Diagnosis"
            sheet['H1']="Medication"
            sheet['I1']="Reference ID"
            sheet['J1']="Ward Name"     
            sheet['K1']="Doctor Name"

            file.save('HMDS_data.xlsx')

        self.fetch_data()

    
        #===========================Functionality Declaration================================
    
    def Add(self):
        if  self.PatientName.get()=="" or self.DoctorName.get()=="":
            messagebox.showerror("Error","All fields are required")
        else:
            file=pathlib.Path('HMDS_data.xlsx')
            if file.exists():
                pass
            else:
                file=Workbook()
                sheet=file.active
                sheet['A1']="Patient Name"
                sheet['B1']="Age"
                sheet['C1']="Patient Address"
                sheet['D1']="Date of Birth"
                sheet['E1']="Date of Visit"
                sheet['F1']="Symptoms"
                sheet['G1']="Diagnosis"
                sheet['H1']="Medication"
                sheet['I1']="Reference ID"
                sheet['J1']="Ward Name"     
                sheet['K1']="Doctor Name"

                file.save('HMDS_data.xlsx')

            file=openpyxl.load_workbook('HMDS_data.xlsx')
            sheet=file.active
            sheet.cell(column=1,row=sheet.max_row+1,value=self.PatientName.get())
            sheet.cell(column=2,row=sheet.max_row,value=self.Age.get())
            sheet.cell(column=3,row=sheet.max_row,value=self.PatientAddress.get())
            sheet.cell(column=4,row=sheet.max_row,value=self.DateofBirth.get())
            sheet.cell(column=5,row=sheet.max_row,value=self.DateofVisit.get())
            sheet.cell(column=6,row=sheet.max_row,value=self.Symptoms.get())
            sheet.cell(column=7,row=sheet.max_row,value=self.Diagnosis.get())
            sheet.cell(column=8,row=sheet.max_row,value=self.Medication.get())
            sheet.cell(column=9,row=sheet.max_row,value=uuid.uuid4().hex[:8])
            sheet.cell(column=10,row=sheet.max_row,value=self.WardName.get())
            sheet.cell(column=11,row=sheet.max_row,value=self.DoctorName.get())

            file.save('HMDS_data.xlsx')
            self.fetch_data()
            self.Clear()
            messagebox.showinfo("Success","Patient Data Added")

    def fetch_data(self):
        file=openpyxl.load_workbook('HMDS_data.xlsx')
        sheet=file.active
        list_values=list(sheet.values)
        rows=sheet.max_row
        if rows!=0:
            self.hospital_table.delete(*self.hospital_table.get_children())
            for value_tuple in list_values[1:]:
                    self.hospital_table.insert('',END,values=value_tuple)
            

    def get_cursor(self,event=""):
        cursor_row=self.hospital_table.focus()
        content=self.hospital_table.item(cursor_row)
        row=content["values"]
        self.PatientName.set(row[0])
        self.Age.set(row[1])
        self.PatientAddress.set(row[2])
        self.DateofBirth.set(row[3])
        self.DateofVisit.set(row[4])
        self.Symptoms.set(row[5])
        self.Diagnosis.set(row[6])
        self.Medication.set(row[7])
        self.WardName.set(row[9])
        self.DoctorName.set(row[10])


    def Update(self):
        
        cursor_iid=self.hospital_table.focus()
        cursor_index=self.hospital_table.index(cursor_iid)
        cursor_index+=2
        file=openpyxl.load_workbook('HMDS_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=cursor_index,value=self.PatientName.get())
        sheet.cell(column=2,row=cursor_index,value=self.Age.get())
        sheet.cell(column=3,row=cursor_index,value=self.PatientAddress.get())
        sheet.cell(column=4,row=cursor_index,value=self.DateofBirth.get())
        sheet.cell(column=5,row=cursor_index,value=self.DateofVisit.get())
        sheet.cell(column=6,row=cursor_index,value=self.Symptoms.get())
        sheet.cell(column=7,row=cursor_index,value=self.Diagnosis.get())
        sheet.cell(column=8,row=cursor_index,value=self.Medication.get())
        sheet.cell(column=10,row=cursor_index,value=self.WardName.get())
        sheet.cell(column=11,row=cursor_index,value=self.DoctorName.get())

        file.save('HMDS_data.xlsx')
        self.fetch_data()
        self.Clear()
        messagebox.showinfo("Success","Patient Data Updated")


    def Delete(self):

        cursor_iid=self.hospital_table.focus()
        cursor_index=self.hospital_table.index(cursor_iid)
        print(cursor_index)
        cursor_index+=2
        file=openpyxl.load_workbook('HMDS_data.xlsx')
        sheet=file.active
        sheet.delete_rows(idx=cursor_index)

        file.save('HMDS_data.xlsx')
        self.fetch_data()
        self.Clear()
        messagebox.showinfo("Success","Patient Data Deleted")


    def Clear(self):

        self.PatientName.set('')
        self.Age.set('')
        self.PatientAddress.set('')
        self.DateofBirth.set('')
        self.DateofVisit.set('')
        self.Symptoms.set('')
        self.Diagnosis.set('')
        self.Medication.set('')
        self.WardName.set('')
        self.DoctorName.set('')




root=Tk()
ob=Hospital(root)
root.mainloop()